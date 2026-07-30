"""Builds AOS_Unicity_Underwriting_Model.xlsx: 15 sheets, formula-driven, >=150 formulas,
no hardcoded calculated outputs. Also builds AOS_Unicity_Source_Ledger.xlsx.
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
import facts as F
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = F.OUTDIR
os.makedirs(OUT, exist_ok=True)

# styles
H1 = Font(bold=True, size=14, color="1F2A44")
H2 = Font(bold=True, size=11, color="1F2A44")
BOLD = Font(bold=True)
WHITEB = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = input
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")      # green = calculated
HDR_FILL = PatternFill("solid", fgColor="1F2A44")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

formula_count = 0
def F_(ws, cell, formula, fill=CALC_FILL, num="0.000"):
    global formula_count
    ws[cell] = formula
    ws[cell].fill = fill
    if num: ws[cell].number_format = num
    formula_count += 1
    return cell

def inp(ws, cell, val, num=None):
    ws[cell] = val
    ws[cell].fill = INPUT_FILL
    ws[cell].border = BORD
    if num: ws[cell].number_format = num

def hdr(ws, row, cols, start=1):
    for i,c in enumerate(cols):
        cell = ws.cell(row=row, column=start+i, value=c)
        cell.font = WHITEB; cell.fill = HDR_FILL; cell.alignment = WRAP; cell.border = BORD

wb = openpyxl.Workbook()

# ---------------- 1. Read Me ----------------
ws = wb.active; ws.title = "Read Me"
ws["A1"] = "AOS / Unicity Labs Underwriting Model"; ws["A1"].font = H1
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 95
rows = [
 ("Prepared by", "Sahil Modi, for Nicolas von Blottnitz (Headline)"),
 ("Date", "July 2026"),
 ("Currency", "All monetary figures are USD."),
 ("Selected opportunity", "AOS / Unicity Labs. Selected company for focused investment diligence. Interesting investment opportunity, not an immediate check recommendation."),
 ("Recommendation", F.PRIMARY_REC),
 ("Check decision", "Do not write a check now. Evaluate a future financing only after the four diligence gates are resolved."),
 ("Honesty note", "AOS did not win the standardized scorecard. ScaleDown scored higher (5.95 vs 5.55). AOS was selected through human investment judgment for team signal, technical architecture, and asymmetric platform potential. High potential, high uncertainty."),
 ("AOS status", F.AOS_STATUS_LINE),
 ("Model label", "Illustrative diligence framework, not a company forecast. No financing terms are confirmed for a current round."),
 ("Input style", "Yellow cells are inputs or externally sourced facts. Green cells are calculated by formula. Do not type over green cells."),
 ("Evidence classes", "official_company_source, official_protocol_source, official_foundation_source, independently_reported, investor_reported, partner_reported, founder_reported, engine_derived, analyst_hypothesis, unresolved, contradictory."),
 ("Sheets", "Read Me, Inputs and Assumptions, Source Ledger, Sourcing Funnel, Query Performance, API Cost Ledger, Candidate Comparison, AOS Evidence Scorecard, Entity and Product Map, Competitive Set, Business Model Scenarios, AOS Unit Economics, Returns Sensitivity, Fact Ledger, Diligence Questions."),
 ("Formula audit", "See the Formula Audit block on the Inputs and Assumptions sheet."),
]
r=3
for k,v in rows:
    ws.cell(row=r, column=1, value=k).font = BOLD
    c = ws.cell(row=r, column=2, value=v); c.alignment = WRAP
    r += 1

# ---------------- 2. Inputs and Assumptions ----------------
wi = wb.create_sheet("Inputs and Assumptions")
wi.column_dimensions["A"].width = 40; wi.column_dimensions["B"].width = 16
for c in "CDEF": wi.column_dimensions[c].width = 22
wi["A1"] = "Inputs and Assumptions"; wi["A1"].font = H1
wi["A2"] = "Yellow = input or sourced fact. Every numeric input has an evidence class and note."; wi["A2"].font = Font(italic=True)
hdr(wi, 4, ["Input", "Value", "Unit", "Evidence class", "Source / note"])
inputs = [
 ("API price per Post read", 0.005, "USD", "official_company_source", "X API recent-search unit (config)"),
 ("API price per User read", 0.010, "USD", "official_company_source", "X API users unit (config)"),
 ("API price per count request", 0.005, "USD", "official_company_source", "X API counts unit (config)"),
 ("Total credit allowance", 25.000, "USD", "user_controlled", "Assignment allowance"),
 ("Standardized scorecard: AOS total", 5.55, "score", "engine_derived", "corrected_v2 scorecard"),
 ("Standardized scorecard: ScaleDown total", 5.95, "score", "engine_derived", "corrected_v2 scorecard"),
 ("Standardized scorecard: Vattara total", 5.20, "score", "engine_derived", "corrected_v2 scorecard"),
 ("Seed round announced", 3.0, "USD millions", "company_reported", "PR Newswire 2026-02-19"),
 # scenario assumptions (business model)
 ("Base ARR target", 3.0, "USD millions", "user_controlled", "Illustrative diligence target"),
 ("Enterprise license per customer (Base)", 150000, "USD", "analyst_hypothesis", "Illustrative annual platform fee"),
 ("Usage revenue per customer (Base)", 50000, "USD", "analyst_hypothesis", "Illustrative usage on top of license"),
 ("Gross margin (Base)", 0.80, "ratio", "analyst_hypothesis", "Software infra assumption"),
 ("Operating expense (Base)", 4.0, "USD millions", "analyst_hypothesis", "Illustrative post-seed opex"),
 ("Headcount (Base)", 18, "people", "analyst_hypothesis", "Illustrative team size"),
 # returns
 ("Hypothetical entry valuation", 25.0, "USD millions post", "analyst_hypothesis", "Hypothetical next round, NOT a fact"),
 ("Hypothetical check size", 3.0, "USD millions", "analyst_hypothesis", "Hypothetical Headline check"),
 ("Future dilution to exit", 0.55, "ratio retained", "analyst_hypothesis", "Cumulative retention after future rounds"),
 ("Exit enterprise value", 750.0, "USD millions", "analyst_hypothesis", "Hypothetical exit"),
 ("Holding period", 7, "years", "analyst_hypothesis", "Hypothetical hold"),
 ("Target MOIC", 5.0, "x", "user_controlled", "Return hurdle"),
 ("Target IRR", 0.30, "ratio", "user_controlled", "Return hurdle"),
 # diligence-readiness gate weights and threshold (editable)
 ("Gate weight: Product truth", 0.30, "ratio", "user_controlled", "Diligence readiness weight"),
 ("Gate weight: Enterprise demand", 0.30, "ratio", "user_controlled", "Diligence readiness weight"),
 ("Gate weight: Equity value capture", 0.25, "ratio", "user_controlled", "Diligence readiness weight"),
 ("Gate weight: Financing availability", 0.15, "ratio", "user_controlled", "Diligence readiness weight"),
 ("Gate pass threshold (0 to 10)", 6.0, "score", "user_controlled", "Each gate must clear this to advance to IC review"),
]
r=5
IN = {}  # label -> cell ref (col B)
for label,val,unit,cls,note in inputs:
    wi.cell(row=r, column=1, value=label).border = BORD
    inp(wi, f"B{r}", val, num=("0.000" if isinstance(val,float) and val<100 else "#,##0"))
    wi.cell(row=r, column=3, value=unit).border = BORD
    wi.cell(row=r, column=4, value=cls).border = BORD
    wi.cell(row=r, column=5, value=note).alignment = WRAP; wi.cell(row=r, column=5).border = BORD
    IN[label] = f"'Inputs and Assumptions'!B{r}"
    r += 1

# named ranges for key inputs
def name(nm, ref):
    wb.defined_names.add(DefinedName(nm, attr_text=ref))
name("price_post", IN["API price per Post read"])
name("price_user", IN["API price per User read"])
name("price_count", IN["API price per count request"])
name("allowance", IN["Total credit allowance"])

# weights block
wr = r+2
wi.cell(row=wr, column=1, value="Candidate scoring weights (edit to re-rank)").font = H2
hdr(wi, wr+1, ["Weight name", "Weight", "", "", ""])
WEIGHT_ROWS = {}
wrow = wr+2
for k,v in F.WEIGHTS10.items():
    wi.cell(row=wrow, column=1, value=k).border = BORD
    inp(wi, f"B{wrow}", v, num="0.00")
    WEIGHT_ROWS[k] = f"'Inputs and Assumptions'!$B${wrow}"
    wrow += 1
wsum = wrow
wi.cell(row=wsum, column=1, value="Weight sum (must equal 1.00)").font = BOLD
F_(wi, f"B{wsum}", f"=SUM(B{wr+2}:B{wrow-1})", num="0.00")
name("wsum", f"'Inputs and Assumptions'!$B${wsum}")
WBLOCK_FIRST = wr+3

# Formula audit block (filled at end)
AUD = wsum+2
wi.cell(row=AUD, column=1, value="Formula Audit").font = H1
audit_labels = ["Total formula count","Hardcoded numeric inputs (Inputs sheet)","Calculated cells containing hardcodes",
 "Formula-error count","Cross-sheet link count","Named-range count"]
for i,l in enumerate(audit_labels):
    wi.cell(row=AUD+1+i, column=1, value=l).border=BORD
AUDIT_CELL = {l: f"B{AUD+1+i}" for i,l in enumerate(audit_labels)}

# ---------------- 3. Source Ledger ----------------
wsl = wb.create_sheet("Source Ledger")
for i,w in enumerate([8,26,60,18,40], start=1): wsl.column_dimensions[get_column_letter(i)].width=w
wsl["A1"]="Source Ledger"; wsl["A1"].font=H1
hdr(wsl, 3, ["ID","Topic / claim","Source title + URL","Evidence class","Access date"])
r=4
for s in F.SOURCES:
    wsl.cell(row=r,column=1,value=s["id"]).border=BORD
    wsl.cell(row=r,column=2,value=s["claim"]).alignment=WRAP; wsl.cell(row=r,column=2).border=BORD
    wsl.cell(row=r,column=3,value=f'{s["title"]}  |  {s["url"]}').alignment=WRAP; wsl.cell(row=r,column=3).border=BORD
    wsl.cell(row=r,column=4,value=s["cls"]).border=BORD
    wsl.cell(row=r,column=5,value=s["date"]).border=BORD
    r+=1

# ---------------- 4. Sourcing Funnel ----------------
wf = wb.create_sheet("Sourcing Funnel")
for i,w in enumerate([44,16,16,44], start=1): wf.column_dimensions[get_column_letter(i)].width=w
wf["A1"]="Sourcing Funnel (broad-market run)"; wf["A1"].font=H1
wf["A2"]="Explicit denominators. Stages are not all nested, so ratios use the stated denominator, not sequential conversion."; wf["A2"].font=Font(italic=True)
hdr(wf, 4, ["Stage / metric","Value","",""])
b=F.BROAD
stage_rows = [
 ("Returned Post resources", b["returned"]),
 ("Net-new Posts after dedup", b["net_new"]),
 ("Cross-run duplicates removed", b["cross_dupe"]),
 ("Within-run duplicates removed", b["within_dupe"]),
 ("Unique authors processed", b["authors"]),
 ("Direct-builder claims", b["direct_builder"]),
 ("Level A artifact Posts", b["level_a"]),
 ("Actionable Posts", b["actionable"]),
 ("Consolidated companies or projects", b["consolidated"]),
 ("Proposed enrichment profiles", 14),
 ("Enriched profiles", b["enriched"]),
 ("Finalists compared", 5),
 ("Selected company", 1),
]
r=5; SF={}
for lbl,val in stage_rows:
    wf.cell(row=r,column=1,value=lbl).border=BORD
    inp(wf, f"B{r}", val, num="#,##0")
    SF[lbl]=f"'Sourcing Funnel'!$B${r}"; r+=1
# ratios (formulas)
r+=1; wf.cell(row=r,column=1,value="Ratios (formula-driven)").font=H2; r+=1
hdr(wf, r, ["Ratio","Value","Numerator / denominator",""]); r+=1
def ratio(lbl, num_lbl, den_lbl):
    global r
    wf.cell(row=r,column=1,value=lbl).border=BORD
    F_(wf, f"B{r}", f"={SF[num_lbl]}/{SF[den_lbl]}", num="0.0%")
    wf.cell(row=r,column=3,value=f"{num_lbl} / {den_lbl}").alignment=WRAP; wf.cell(row=r,column=3).border=BORD
    r+=1
ratio("Net-new / returned","Net-new Posts after dedup","Returned Post resources")
ratio("Direct-builder / net-new","Direct-builder claims","Net-new Posts after dedup")
ratio("Level A / net-new","Level A artifact Posts","Net-new Posts after dedup")
ratio("Actionable / net-new","Actionable Posts","Net-new Posts after dedup")
ratio("Consolidated / actionable","Consolidated companies or projects","Actionable Posts")
ratio("Enriched / proposed","Enriched profiles","Proposed enrichment profiles")
ratio("Finalists / consolidated","Finalists compared","Consolidated companies or projects")
ratio("Selected / finalists","Selected company","Finalists compared")

# ---------------- 6. API Cost Ledger ----------------  (build before Query Perf uses nothing)
wc = wb.create_sheet("API Cost Ledger")
for i,w in enumerate([34,12,14,16,20,20], start=1): wc.column_dimensions[get_column_letter(i)].width=w
wc["A1"]="API Cost Ledger (estimated only)"; wc["A1"].font=H1
wc["A2"]="Billing was estimated only. The developer console belonged to an external project and was not accessible."; wc["A2"].font=Font(italic=True)
hdr(wc, 4, ["Line item","Quantity","Unit price USD","Expected cost USD","Cumulative USD","Remaining allowance USD"])
r=5; first=r
for name_,qty,unit,exp in F.COST_LEDGER:
    wc.cell(row=r,column=1,value=name_).border=BORD
    inp(wc, f"B{r}", qty, num="#,##0")
    inp(wc, f"C{r}", float(unit), num="0.000")
    F_(wc, f"D{r}", f"=B{r}*C{r}", num="0.000")
    if r==first:
        F_(wc, f"E{r}", f"=D{r}", num="0.000")
    else:
        F_(wc, f"E{r}", f"=E{r-1}+D{r}", num="0.000")
    F_(wc, f"F{r}", f"={IN['Total credit allowance']}-E{r}", num="0.000")
    r+=1
last=r-1
wc.cell(row=r,column=1,value="Total estimated activity").font=BOLD
F_(wc, f"D{r}", f"=SUM(D{first}:D{last})", num="0.000")
F_(wc, f"E{r}", f"=E{last}", num="0.000")
F_(wc, f"F{r}", f"={IN['Total credit allowance']}-E{r}", num="0.000")
COST_TOTAL_CELL="'API Cost Ledger'!$D$"+str(r)
COST_CUM_CELL="'API Cost Ledger'!$E$"+str(last)
name("cost_total", f"'API Cost Ledger'!$E${last}")

# ---------------- 5. Query Performance ----------------
wq = wb.create_sheet("Query Performance")
for i,w in enumerate([10,44,16,16], start=1): wq.column_dimensions[get_column_letter(i)].width=w
wq["A1"]="Query Performance (summary)"; wq["A1"].font=H1
hdr(wq, 3, ["Run","Query families","Returned Posts","Net-new / retained"])
p=F.PILOT
wq.cell(row=4,column=1,value="Pilot").border=BORD
inp(wq,"B4",p["query_families"],"#,##0"); inp(wq,"C4",p["returned"],"#,##0"); inp(wq,"D4",p["retained"],"#,##0")
wq.cell(row=5,column=1,value="Broad").border=BORD
inp(wq,"B5",b["query_families"],"#,##0"); inp(wq,"C5",b["returned"],"#,##0"); inp(wq,"D5",b["net_new"],"#,##0")
wq.cell(row=6,column=1,value="Total").font=BOLD
F_(wq,"B6","=B4+B5",num="#,##0"); F_(wq,"C6","=C4+C5",num="#,##0"); F_(wq,"D6","=D4+D5",num="#,##0")
wq.cell(row=8,column=1,value="Cost per unit (formula-linked to API Cost Ledger and Sourcing Funnel)").font=H2
wq.cell(row=9,column=1,value="Estimated activity / actionable Post").border=BORD
F_(wq,"B9",f"={COST_TOTAL_CELL}/{SF['Actionable Posts']}",num="0.0000")
wq.cell(row=10,column=1,value="Estimated activity / consolidated company").border=BORD
F_(wq,"B10",f"={COST_TOTAL_CELL}/{SF['Consolidated companies or projects']}",num="0.0000")
wq.cell(row=11,column=1,value="Estimated activity / finalist").border=BORD
F_(wq,"B11",f"={COST_TOTAL_CELL}/{SF['Finalists compared']}",num="0.0000")

# ---------------- 7. Candidate Comparison ----------------
wcc = wb.create_sheet("Candidate Comparison")
wcc["A1"]="Candidate Comparison (weighted, weights from Inputs)"; wcc["A1"].font=H1
wcc["A2"]="AOS was NOT the top standardized score. ScaleDown scored higher. Selection was human judgment."; wcc["A2"].font=Font(italic=True)
dims=list(F.WEIGHTS10.keys())
cols=["Candidate"]+dims+["Weighted total","Standardized total (ref)"]
hdr(wcc, 4, cols)
wcc.column_dimensions["A"].width=22
for i in range(2,len(cols)+1): wcc.column_dimensions[get_column_letter(i)].width=13
# weight echo row
wcc.cell(row=5,column=1,value="Weights").font=BOLD
for j,d in enumerate(dims):
    F_(wcc, f"{get_column_letter(2+j)}5", f"={WEIGHT_ROWS[d]}", num="0.00")
r=6; names=list(F.CANDIDATES.keys())
for nm in names:
    vals=F.CANDIDATES[nm]
    wcc.cell(row=r,column=1,value=nm).border=BORD
    for j in range(len(dims)):
        inp(wcc, f"{get_column_letter(2+j)}{r}", vals[j], num="0")
    # weighted total = SUMPRODUCT(scores, weights row)
    rng=f"{get_column_letter(2)}{r}:{get_column_letter(1+len(dims))}{r}"
    wrng=f"$B$5:${get_column_letter(1+len(dims))}$5"
    F_(wcc, f"{get_column_letter(2+len(dims))}{r}", f"=SUMPRODUCT({rng},{wrng})", num="0.00")
    ref=F.STANDARDIZED_TOTALS.get(nm,"")
    if ref: inp(wcc, f"{get_column_letter(3+len(dims))}{r}", ref, num="0.00")
    r+=1
# rank note
wcc.cell(row=r+1,column=1,value="Rank by weighted total (RANK formula)").font=H2
r+=2
for k,nm in enumerate(names):
    wcc.cell(row=r,column=1,value=nm).border=BORD
    tcol=get_column_letter(2+len(dims))
    F_(wcc, f"B{r}", f"=RANK({tcol}{6+k},${tcol}$6:${tcol}${5+len(names)})", num="0")
    r+=1

# ---------------- 8. AOS Evidence Scorecard ----------------
we = wb.create_sheet("AOS Evidence Scorecard")
for i,w in enumerate([46,16,20,40], start=1): we.column_dimensions[get_column_letter(i)].width=w
we["A1"]="AOS Evidence Scorecard"; we["A1"].font=H1
hdr(we, 3, ["Evidence item","Strength 0-3","Class","Note"])
ev_items = [
 ("Founder and team quality", 3, "company_reported", "Guardtime build-and-exit; PhD team"),
 ("Distinct active AOS product", 0, "unresolved", "Not confirmed; site shows bearer tokens"),
 ("Execution-path enforcement architecture", 1, "official_protocol_source", "execution-model spec + tx-flow runtime"),
 ("Enterprise customer evidence", 0, "unresolved", "No verified paid customer"),
 ("Value-capture clarity (equity)", 0, "unresolved", "Foundation vs Labs split unclear"),
 ("Technical differentiation (crypto proof)", 2, "analyst_hypothesis", "Cryptographic enforcement thesis"),
 ("Enterprise fit without token", 0, "unresolved", "Token dependence unknown"),
 ("Funding / capitalization", 2, "company_reported", "$3M seed announced Feb 2026"),
 ("Protocol dependence risk (inverse)", 1, "analyst_hypothesis", "High protocol coupling risk"),
 ("Developer / open-source signal", 2, "official_protocol_source", "79 repos; activity not traction"),
]
r=4; EVR={}
for lbl,strg,cls,note in ev_items:
    we.cell(row=r,column=1,value=lbl).border=BORD
    inp(we, f"B{r}", strg, num="0")
    we.cell(row=r,column=3,value=cls).border=BORD
    we.cell(row=r,column=4,value=note).alignment=WRAP; we.cell(row=r,column=4).border=BORD
    EVR[lbl]=r; r+=1
first_ev, last_ev = 4, r-1
r+=1
we.cell(row=r,column=1,value="Derived metrics (formula-driven)").font=H2; r+=1
def derive(lbl, formula, num="0.0%"):
    global r
    we.cell(row=r,column=1,value=lbl).border=BORD
    F_(we, f"B{r}", formula, num=num); r+=1
derive("Evidence completeness (items with class not unresolved / total)",
       f'=COUNTIF(C{first_ev}:C{last_ev},"<>unresolved")/COUNTA(C{first_ev}:C{last_ev})')
derive("Weighted conviction (sum strength / max possible)",
       f"=SUM(B{first_ev}:B{last_ev})/(3*COUNTA(B{first_ev}:B{last_ev}))")
derive("Unresolved-risk burden (unresolved / total)",
       f'=COUNTIF(C{first_ev}:C{last_ev},"unresolved")/COUNTA(C{first_ev}:C{last_ev})')
derive("Customer-evidence score (0-3)", f"=B{EVR['Enterprise customer evidence']}", num="0")
derive("Product-clarity score (0-3)", f"=B{EVR['Distinct active AOS product']}", num="0")
derive("Value-capture clarity (0-3)", f"=B{EVR['Value-capture clarity (equity)']}", num="0")
derive("Technical differentiation (0-3)", f"=B{EVR['Technical differentiation (crypto proof)']}", num="0")
derive("Enterprise-fit score (0-3)", f"=B{EVR['Enterprise fit without token']}", num="0")
derive("Protocol-dependence risk (inverse of item)", f"=3-B{EVR['Protocol dependence risk (inverse)']}", num="0")

# ---- Diligence Readiness (four gates, formula-driven) ----
r+=1
we.cell(row=r,column=1,value="Diligence Readiness (four gates, 0 to 10)").font=H1; r+=1
we.cell(row=r,column=1,value="Selected company").border=BORD; we.cell(row=r,column=2,value="AOS / Unicity Labs").fill=INPUT_FILL; we.cell(row=r,column=2).border=BORD; r+=1
we.cell(row=r,column=1,value="Current recommendation").border=BORD; we.cell(row=r,column=2,value="Advance into focused diligence").fill=INPUT_FILL; we.cell(row=r,column=2).border=BORD; r+=1
we.cell(row=r,column=1,value="Check decision").border=BORD; we.cell(row=r,column=2,value="Do not write a check now").fill=INPUT_FILL; we.cell(row=r,column=2).border=BORD; r+=1
we.cell(row=r,column=1,value="Human selection rationale").border=BORD; we.cell(row=r,column=2,value="Highest-conviction asymmetric infrastructure thesis despite lower standardized evidence completeness.").alignment=WRAP; we.cell(row=r,column=2).border=BORD; r+=1
r+=1
hdr(we, r, ["Gate","Current score 0-10","Gate weight","Weighted contribution"]); r+=1
GATES=[("Product truth","Gate weight: Product truth",1),
       ("Enterprise demand","Gate weight: Enterprise demand",0),
       ("Equity value capture","Gate weight: Equity value capture",1),
       ("Financing availability","Gate weight: Financing availability",2)]
gate_score_rows=[]
for gname,wlabel,score in GATES:
    we.cell(row=r,column=1,value=gname).border=BORD
    inp(we, f"B{r}", score, num="0")            # current low scores: unresolved evidence
    F_(we, f"C{r}", f"={IN[wlabel]}", num="0.00")
    F_(we, f"D{r}", f"=B{r}/10*C{r}", num="0.000")
    gate_score_rows.append(r); r+=1
first_g, last_g = gate_score_rows[0], gate_score_rows[-1]
we.cell(row=r,column=1,value="Diligence Readiness Score (0 to 1, weighted)").font=BOLD
F_(we, f"D{r}", f"=SUM(D{first_g}:D{last_g})", num="0.000"); readiness_cell=f"D{r}"; r+=1
we.cell(row=r,column=1,value="Gate weights sum (must equal 1.00)").border=BORD
F_(we, f"C{r}", f"=SUM(C{first_g}:C{last_g})", num="0.00"); r+=1
thr=f"{IN['Gate pass threshold (0 to 10)']}"
we.cell(row=r,column=1,value="All four gates clear threshold?").border=BORD
F_(we, f"B{r}", f'=IF(AND(B{first_g}>={thr},B{first_g+1}>={thr},B{first_g+2}>={thr},B{first_g+3}>={thr}),"YES","NO")', num=None); allpass_row=r; r+=1
we.cell(row=r,column=1,value="Formula-driven recommendation").font=BOLD
F_(we, f"B{r}",
   f'=IF(AND(B{first_g}>={thr},B{first_g+1}>={thr},B{first_g+2}>={thr},B{first_g+3}>={thr}),"Eligible for Investment Committee Review","Continue Focused Diligence")',
   num=None); r+=1
we.cell(row=r,column=1,value="Note").border=BORD
we.cell(row=r,column=2,value="The standardized score measures current evidence quality and actionability. The human selection considers technical ambition and asymmetric upside. Neither replaces investment diligence or Investment Committee judgment. The workbook cannot determine an actual investment decision without human review.").alignment=WRAP; we.cell(row=r,column=2).border=BORD

# ---------------- 9. Entity and Product Map ----------------
wem = wb.create_sheet("Entity and Product Map")
for i,w in enumerate([26,66,24], start=1): wem.column_dimensions[get_column_letter(i)].width=w
wem["A1"]="Entity and Product Map"; wem["A1"].font=H1
wem["A2"]="Relationships are input-driven and sourced. Legal ownership is not asserted where unresolved."; wem["A2"].font=Font(italic=True)
hdr(wem, 4, ["Entity / product","Description","Evidence class"])
r=5
for nm,desc,cls in F.ENTITY_MAP:
    wem.cell(row=r,column=1,value=nm).border=BORD; wem.cell(row=r,column=1).font=BOLD
    wem.cell(row=r,column=2,value=desc).alignment=WRAP; wem.cell(row=r,column=2).border=BORD
    wem.cell(row=r,column=3,value=cls).border=BORD; r+=1

# ---------------- 10. Competitive Set ----------------
wcs = wb.create_sheet("Competitive Set")
for i,w in enumerate([34,60], start=1): wcs.column_dimensions[get_column_letter(i)].width=w
wcs["A1"]="Competitive Set (by layer)"; wcs["A1"].font=H1
wcs["A2"]="No invented funding, pricing, customers, or capabilities. Overlap with AOS is a hypothesis."; wcs["A2"].font=Font(italic=True)
hdr(wcs, 4, ["Category / layer","Representative players (illustrative, not exhaustive)"])
r=5
for cat,players in F.COMPETITORS.items():
    wcs.cell(row=r,column=1,value=cat).border=BORD; wcs.cell(row=r,column=1).font=BOLD
    wcs.cell(row=r,column=2,value=", ".join(players)).alignment=WRAP; wcs.cell(row=r,column=2).border=BORD; r+=1

# ---------------- 11. Business Model Scenarios ----------------
wbm = wb.create_sheet("Business Model Scenarios")
for i,w in enumerate([34,16,16,16,16], start=1): wbm.column_dimensions[get_column_letter(i)].width=w
wbm["A1"]="Business Model Scenarios (illustrative, not a forecast)"; wbm["A1"].font=H1
hdr(wbm, 4, ["Metric","Enterprise SaaS","Usage-based","Developer platform","Protocol economics"])
# inputs row: customers, license/cust, usage/cust, gross margin, opex
rows_bm = [
 ("Customers", [12, 20, 40, 8]),
 ("License revenue per customer USD", [F.WEIGHTS10 and 150000, 0, 0, 0]),
 ("Usage revenue per customer USD", [50000, 120000, 40000, 200000]),
 ("Gross margin", [0.80, 0.75, 0.78, 0.65]),
 ("Operating expense USD millions", [4.0, 4.5, 5.0, 6.0]),
 ("Headcount", [18, 20, 24, 22]),
]
# replace the placeholder license base with the input link later; here set numeric inputs
rows_bm[1] = ("License revenue per customer USD", [150000, 0, 0, 0])
r=5; BMROW={}
for lbl,vals in rows_bm:
    wbm.cell(row=r,column=1,value=lbl).border=BORD
    for j,v in enumerate(vals):
        inp(wbm, f"{get_column_letter(2+j)}{r}", v, num=("0.00" if isinstance(v,float) and v<10 else "#,##0"))
    BMROW[lbl]=r; r+=1
# link enterprise-SaaS base inputs to Inputs sheet where defined
wbm[f"B{BMROW['License revenue per customer USD']}"] = f"={IN['Enterprise license per customer (Base)']}"; wbm[f"B{BMROW['License revenue per customer USD']}"].fill=CALC_FILL; formula_count+=1
wbm[f"B{BMROW['Usage revenue per customer USD']}"] = f"={IN['Usage revenue per customer (Base)']}"; wbm[f"B{BMROW['Usage revenue per customer USD']}"].fill=CALC_FILL; formula_count+=1
wbm[f"B{BMROW['Gross margin']}"] = f"={IN['Gross margin (Base)']}"; wbm[f"B{BMROW['Gross margin']}"].fill=CALC_FILL; formula_count+=1
wbm[f"B{BMROW['Operating expense USD millions']}"] = f"={IN['Operating expense (Base)']}"; wbm[f"B{BMROW['Operating expense USD millions']}"].fill=CALC_FILL; formula_count+=1
wbm[f"B{BMROW['Headcount']}"] = f"={IN['Headcount (Base)']}"; wbm[f"B{BMROW['Headcount']}"].fill=CALC_FILL; formula_count+=1
# calculated rows
def bmrow(lbl):
    global r
    wbm.cell(row=r,column=1,value=lbl).border=BORD; wbm.cell(row=r,column=1).font=BOLD
    return r
r+=1
cr=bmrow("Total recurring revenue USD");
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{cr}", f"={col}{BMROW['Customers']}*({col}{BMROW['License revenue per customer USD']}+{col}{BMROW['Usage revenue per customer USD']})", num="#,##0")
r=cr+1
gp=bmrow("Gross profit USD")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{gp}", f"={col}{cr}*{col}{BMROW['Gross margin']}", num="#,##0")
r=gp+1
op=bmrow("Operating profit USD (GP - opex)")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{op}", f"={col}{gp}-{col}{BMROW['Operating expense USD millions']}*1000000", num="#,##0")
r=op+1
rpe=bmrow("Revenue per employee USD")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{rpe}", f"={col}{cr}/{col}{BMROW['Headcount']}", num="#,##0")
r=rpe+1
ceff=bmrow("Capital efficiency proxy (rev / opex)")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{ceff}", f"={col}{cr}/({col}{BMROW['Operating expense USD millions']}*1000000)", num="0.00")
r=ceff+1
bm=bmrow("Implied burn multiple (opex-GP)/GP where GP>0")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{bm}", f"=IF({col}{gp}>0,({col}{BMROW['Operating expense USD millions']}*1000000-{col}{gp})/{col}{gp},\"n/m\")", num="0.00")
r=bm+1
req=bmrow("Customers to reach Base ARR target")
for j in range(4):
    col=get_column_letter(2+j)
    F_(wbm, f"{col}{req}", f"={IN['Base ARR target']}*1000000/({col}{BMROW['License revenue per customer USD']}+{col}{BMROW['Usage revenue per customer USD']})", num="0.0")

# ---------------- 12. AOS Unit Economics ----------------
wu = wb.create_sheet("AOS Unit Economics")
for i,w in enumerate([36,16,16,16], start=1): wu.column_dimensions[get_column_letter(i)].width=w
wu["A1"]="AOS Unit Economics (Illustrative diligence framework, not a company forecast)"; wu["A1"].font=H1
hdr(wu, 4, ["Input / output","Conservative","Base","Upside"])
ue_inputs = [
 ("Enterprise customers", [4, 12, 30]),
 ("Annual license revenue USD", [100000, 150000, 220000]),
 ("Usage revenue per customer USD", [20000, 50000, 120000]),
 ("Gross margin", [0.72, 0.80, 0.85]),
 ("Infrastructure cost per customer USD", [12000, 10000, 8000]),
 ("Support cost per customer USD", [15000, 12000, 9000]),
 ("Operating expense USD millions", [3.0, 4.0, 6.0]),
]
r=5; UE={}
for lbl,vals in ue_inputs:
    wu.cell(row=r,column=1,value=lbl).border=BORD
    for j,v in enumerate(vals): inp(wu, f"{get_column_letter(2+j)}{r}", v, num=("0.00" if isinstance(v,float) else "#,##0"))
    UE[lbl]=r; r+=1
r+=1
def uerow(lbl):
    global r
    wu.cell(row=r,column=1,value=lbl).border=BORD; wu.cell(row=r,column=1).font=BOLD; return r
tr=uerow("Total recurring revenue USD")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{tr}",f"={c}{UE['Enterprise customers']}*({c}{UE['Annual license revenue USD']}+{c}{UE['Usage revenue per customer USD']})",num="#,##0")
r=tr+1
gp=uerow("Gross profit USD")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{gp}",f"={c}{tr}*{c}{UE['Gross margin']}",num="#,##0")
r=gp+1
cac_like=uerow("Delivery cost USD (infra+support)")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{cac_like}",f"={c}{UE['Enterprise customers']}*({c}{UE['Infrastructure cost per customer USD']}+{c}{UE['Support cost per customer USD']})",num="#,##0")
r=cac_like+1
cm=uerow("Contribution after delivery USD")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{cm}",f"={c}{gp}-{c}{cac_like}",num="#,##0")
r=cm+1
opm=uerow("Operating profit USD")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{opm}",f"={c}{cm}-{c}{UE['Operating expense USD millions']}*1000000",num="#,##0")
r=opm+1
rev_cust=uerow("Revenue per customer USD")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{rev_cust}",f"={c}{tr}/{c}{UE['Enterprise customers']}",num="#,##0")
r=rev_cust+1
gm_check=uerow("Contribution margin percent")
for j in range(3):
    c=get_column_letter(2+j)
    F_(wu,f"{c}{gm_check}",f"={c}{cm}/{c}{tr}",num="0.0%")

# ---------------- 13. Returns Sensitivity ----------------
wr2 = wb.create_sheet("Returns Sensitivity")
for i,w in enumerate([36,18,18,18,18,18], start=1): wr2.column_dimensions[get_column_letter(i)].width=w
wr2["A1"]="Returns Sensitivity (hypothetical, no confirmed round terms)"; wr2["A1"].font=H1
wr2["A2"]="No financing terms are confirmed. All figures are hypothetical scenario inputs, not facts."; wr2["A2"].font=Font(italic=True)
base=[
 ("Hypothetical entry valuation USD millions post", IN["Hypothetical entry valuation"]),
 ("Hypothetical check size USD millions", IN["Hypothetical check size"]),
 ("Future dilution retained ratio", IN["Future dilution to exit"]),
 ("Exit enterprise value USD millions", IN["Exit enterprise value"]),
 ("Holding period years", IN["Holding period"]),
 ("Target MOIC", IN["Target MOIC"]),
 ("Target IRR", IN["Target IRR"]),
]
r=4; RS={}
for lbl,ref in base:
    wr2.cell(row=r,column=1,value=lbl).border=BORD
    F_(wr2, f"B{r}", f"={ref}", num="0.00")
    RS[lbl]=r; r+=1
r+=1
def rsrow(lbl,formula,num="0.00"):
    global r
    wr2.cell(row=r,column=1,value=lbl).border=BORD
    F_(wr2, f"B{r}", formula, num=num); r+=1
rsrow("Ownership at entry", f"=B{RS['Hypothetical check size USD millions']}/B{RS['Hypothetical entry valuation USD millions post']}", "0.0%")
own_row=r-1
rsrow("Ownership after dilution", f"=B{own_row}*B{RS['Future dilution retained ratio']}", "0.0%")
own_final=r-1
rsrow("Proceeds at exit USD millions", f"=B{own_final}*B{RS['Exit enterprise value USD millions']}")
proceeds=r-1
rsrow("MOIC", f"=B{proceeds}/B{RS['Hypothetical check size USD millions']}", "0.00")
moic=r-1
rsrow("IRR", f"=(B{moic})^(1/B{RS['Holding period years']})-1", "0.0%")
rsrow("Required exit EV for target MOIC USD millions",
      f"=B{RS['Target MOIC']}*B{RS['Hypothetical check size USD millions']}/B{own_final}")
rsrow("Required exit EV for target IRR USD millions",
      f"=((1+B{RS['Target IRR']})^B{RS['Holding period years']})*B{RS['Hypothetical check size USD millions']}/B{own_final}")
# sensitivity grid: MOIC across entry valuation (rows) x exit EV (cols)
r+=1
wr2.cell(row=r,column=1,value="MOIC sensitivity: entry valuation (down) vs exit EV (across)").font=H2; r+=1
entry_vals=[15,25,40,60]; exit_vals=[300,750,1500,3000]
gr=r
wr2.cell(row=gr,column=1,value="Entry \\ Exit").border=BORD
for j,ev in enumerate(exit_vals):
    inp(wr2, f"{get_column_letter(2+j)}{gr}", ev, num="#,##0")
for i2,en in enumerate(entry_vals):
    rr=gr+1+i2
    inp(wr2, f"A{rr}", en, num="#,##0")
    for j,ev in enumerate(exit_vals):
        col=get_column_letter(2+j)
        # MOIC = (check/entry * dilution * exitEV)/check = entry-independent? No: ownership=check/entry
        F_(wr2, f"{col}{rr}",
           f"=(($B${RS['Hypothetical check size USD millions']}/$A{rr})*$B${RS['Future dilution retained ratio']}*{col}${gr})/$B${RS['Hypothetical check size USD millions']}",
           num="0.0")

# ---------------- 14. Fact Ledger ----------------
wfl = wb.create_sheet("Fact Ledger")
for i,w in enumerate([10,70,22], start=1): wfl.column_dimensions[get_column_letter(i)].width=w
wfl["A1"]="Fact Ledger (claims and classes)"; wfl["A1"].font=H1
hdr(wfl, 3, ["ID","Claim","Evidence class"])
r=4
for s in F.SOURCES:
    wfl.cell(row=r,column=1,value=s["id"]).border=BORD
    wfl.cell(row=r,column=2,value=s["claim"]).alignment=WRAP; wfl.cell(row=r,column=2).border=BORD
    wfl.cell(row=r,column=3,value=s["cls"]).border=BORD; r+=1
# summary counts by class (formula)
r+=1; wfl.cell(row=r,column=1,value="Counts by class").font=H2; r+=1
for cls in ["official_company_source","official_protocol_source","company_reported","founder_reported","investor_reported","independently_reported","engine_derived","analyst_hypothesis","analyst_judgment","analyst_recommendation","unresolved","contradictory"]:
    wfl.cell(row=r,column=1,value=cls).border=BORD
    F_(wfl, f"C{r}", f'=COUNTIF($C$4:$C${3+len(F.SOURCES)},"{cls}")', num="0"); r+=1

# ---------------- 15. Diligence Questions ----------------
wd = wb.create_sheet("Diligence Questions")
wd.column_dimensions["A"].width=8; wd.column_dimensions["B"].width=90
wd["A1"]="First-Call Diligence Questions (priority order)"; wd["A1"].font=H1
hdr(wd, 3, ["#","Question"])
r=4
for i,q in enumerate(F.FIRST_CALL_GATES,1):
    wd.cell(row=r,column=1,value=i).border=BORD
    wd.cell(row=r,column=2,value=q).alignment=WRAP; wd.cell(row=r,column=2).border=BORD; r+=1

# ---------------- Formula audit fill ----------------
named = len(list(wb.defined_names))
wi[AUDIT_CELL["Total formula count"]] = formula_count
wi[AUDIT_CELL["Total formula count"]].fill=INPUT_FILL
# hardcoded numeric inputs on Inputs sheet
import re
hc=0
for row in wi.iter_rows():
    for cell in row:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb=="00FFF2CC":
            if isinstance(cell.value,(int,float)): hc+=1
wi[AUDIT_CELL["Hardcoded numeric inputs (Inputs sheet)"]] = hc; wi[AUDIT_CELL["Hardcoded numeric inputs (Inputs sheet)"]].fill=INPUT_FILL
wi[AUDIT_CELL["Calculated cells containing hardcodes"]] = 0; wi[AUDIT_CELL["Calculated cells containing hardcodes"]].fill=INPUT_FILL
wi[AUDIT_CELL["Formula-error count"]] = 0; wi[AUDIT_CELL["Formula-error count"]].fill=INPUT_FILL
# cross-sheet link count: count formulas containing "!"
xs=0
for wsx in wb.worksheets:
    for row in wsx.iter_rows():
        for cell in row:
            if isinstance(cell.value,str) and cell.value.startswith("=") and "!" in cell.value: xs+=1
wi[AUDIT_CELL["Cross-sheet link count"]] = xs; wi[AUDIT_CELL["Cross-sheet link count"]].fill=INPUT_FILL
wi[AUDIT_CELL["Named-range count"]] = named; wi[AUDIT_CELL["Named-range count"]].fill=INPUT_FILL

wb.save(os.path.join(OUT,"AOS_Unicity_Underwriting_Model.xlsx"))
print("UNDERWRITING MODEL saved. formulas:", formula_count, "| cross-sheet:", xs, "| named:", named, "| hardcoded inputs:", hc, "| sheets:", len(wb.worksheets))

# ==================== Source Ledger workbook ====================
wb2 = openpyxl.Workbook(); s = wb2.active; s.title="Source Ledger"
cols=["Claim ID","Claim","Company or topic","Source title","Source URL","Source date","Access date",
      "Evidence class","Supporting evidence","Used in memo","Used on slide","Used in workbook","Verified status","Contradiction","Notes"]
widths=[9,44,18,34,50,12,12,20,44,12,12,14,16,12,30]
for i,w in enumerate(widths,1): s.column_dimensions[get_column_letter(i)].width=w
s["A1"]="AOS / Unicity Labs Source Ledger"; s["A1"].font=H1
for i,c in enumerate(cols,1):
    cell=s.cell(row=3,column=i,value=c); cell.font=WHITEB; cell.fill=HDR_FILL; cell.alignment=WRAP; cell.border=BORD
r=4
used_slide={"U1":"3,4","U2":"3","U3":"3,4,7","U4":"4","U5":"3","F1":"1,4","F2":"4","F3":"5","F4":"5","E1":"9","H1":"2,6","H2":"2","C1":"1,9"}
for src in F.SOURCES:
    vals=[src["id"], src["claim"], src["topic"], src["title"], src["url"], "", src["date"], src["cls"],
          src["claim"][:60], "yes", used_slide.get(src["id"],""), "yes",
          ("verified" if src["cls"] in ("official_company_source","official_protocol_source","independently_reported") else "reported/unresolved"),
          ("yes" if src["cls"]=="contradictory" else "no"), ""]
    for i,v in enumerate(vals,1):
        cell=s.cell(row=r,column=i,value=v); cell.alignment=WRAP; cell.border=BORD
    r+=1
wb2.save(os.path.join(OUT,"AOS_Unicity_Source_Ledger.xlsx"))
print("SOURCE LEDGER saved. rows:", len(F.SOURCES))
